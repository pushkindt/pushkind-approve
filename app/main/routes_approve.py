from app import db
from flask_login import current_user, login_required
from app.main import bp
from app.models import User, UserRoles, Ecwid, OrderApproval, OrderEvent, EventType, OrderStatus
from app.models import Project, Category, Order, OrderCategory, Site, AppSettings, OrderPosition
from app.models import IncomeStatement, CashflowStatement, OrderLimit
from flask import render_template, redirect, url_for, flash, Response, request
from app.main.forms import LeaveCommentForm, OrderApprovalForm, ChangeQuantityForm, InitiativeForm
from app.main.forms import ApproverForm, SplitOrderForm
from datetime import datetime, timezone, date, timedelta
from app.ecwid import EcwidAPIException
from app.main.utils import role_required, ecwid_required, role_forbidden, SendEmailNotification, SendEmail1C, GetNewOrderNumber

from openpyxl import load_workbook
from copy import copy
from openpyxl.writer.excel import save_virtual_workbook
from sqlalchemy.orm.attributes import flag_modified

'''
################################################################################
Approve page
################################################################################
'''


@bp.app_template_filter()
def intersect(a, b):
    return set(a).intersection(set(b))


def GetOrder(order_id):
    order = Order.query.filter_by(id=order_id, hub_id=current_user.hub_id)
    if current_user.role == UserRoles.initiative:
        order = order.filter_by(initiative_id=current_user.id)
    elif current_user.role in [UserRoles.purchaser, UserRoles.validator]:
        order = order.join(OrderCategory).filter(
            OrderCategory.category_id.in_([cat.id for cat in current_user.categories]))
        order = order.join(Site).filter(Site.project_id.in_(
            [p.id for p in current_user.projects]))
    order = order.first()
    return order


@bp.route('/orders/<order_id>')
@login_required
@role_forbidden([UserRoles.default])
@ecwid_required
def ShowOrder(order_id):

    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))

    approval_form = OrderApprovalForm()
    quantity_form = ChangeQuantityForm()
    comment_form = LeaveCommentForm()
    initiative_form = InitiativeForm()
    approver_form = ApproverForm()

    incomes = IncomeStatement.query.filter(IncomeStatement.hub_id == current_user.hub_id).order_by(IncomeStatement.name).all()
    cashflows = CashflowStatement.query.filter(CashflowStatement.hub_id == current_user.hub_id).order_by(CashflowStatement.name).all()

    projects = Project.query
    if current_user.role != UserRoles.admin:
        projects = projects.filter_by(enabled=True)
    projects = projects.filter_by(hub_id=current_user.hub_id)
    projects = projects.order_by(Project.name).all()

    categories = Category.query.filter(
        Category.hub_id == current_user.hub_id).all()

    initiative_form.categories.choices = [(c.id, c.name) for c in categories]
    initiative_form.categories.default = order.categories_list

    approver_form.income_statement.choices = [(i.id, i.name) for i in incomes]
    approver_form.cashflow_statement.choices = [(c.id, c.name) for c in cashflows]
    
    if order.income_statement is None:
        approver_form.income_statement.choices.append((0, 'Выберите БДР...'))
        approver_form.income_statement.default = 0
    else:
        approver_form.income_statement.default = order.income_statement.id
        
    if order.cashflow_statement is None:
        approver_form.cashflow_statement.choices.append((0, 'Выберите БДДС...'))
        approver_form.cashflow_statement.default = 0
    else:
        approver_form.cashflow_statement.default = order.cashflow_statement.id
        
    approver_form.process()

    initiative_form.project.choices = [(p.id, p.name) for p in projects]
    if order.site is None:
        initiative_form.project.choices.append((0, 'Выберите проект...'))
        initiative_form.project.default = 0
        initiative_form.site.choices = [(0, 'Выберите объект...')]
        initiative_form.site.default = 0
    else:
        initiative_form.project.default = order.site.project_id
        initiative_form.site.choices = [
            (s.id, s.name) for s in order.site.project.sites]
        initiative_form.site.default = order.site_id
    initiative_form.process()
    
    split_form = SplitOrderForm()

    return render_template('approve.html',
                           order=order,
                           projects=projects,
                           comment_form=comment_form,
                           approval_form=approval_form,
                           quantity_form=quantity_form,
                           initiative_form=initiative_form,
                           approver_form=approver_form,
                           split_form=split_form)



@bp.route('/orders/split/<order_id>', methods=['POST'])
@login_required
@role_required([UserRoles.admin, UserRoles.initiative, UserRoles.purchaser])
@ecwid_required
def SplitOrder(order_id):

    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))

    if len(order.children) > 0:
        flash('Нельзя разделять заявки, которые были объединены или разделены.')
        return redirect(url_for('main.ShowIndex'))

    form = SplitOrderForm()
    if form.validate_on_submit():
        product_ids = form.products.data
        if not isinstance(product_ids, list) or len(product_ids) == 0:
            flash('Некорректный список позиции.')
            return redirect(url_for('main.ShowOrder', order_id=order_id))

       
        product_lists = [[], []]
        
        for product in order.products:
            if str(product['id']) in product_ids:
                product_lists[0].append(product)
            else:
                product_lists[1].append(product)

        if len(product_lists[0]) == 0 or len(product_lists[1]) == 0:
            flash('Некорректный список позиции.')
            return redirect(url_for('main.ShowOrder', order_id=order_id))

        message_flash = f'заявка разделена на заявки'

        for product_list in product_lists:
            new_order_id = GetNewOrderNumber() 
            message_flash += f' {new_order_id}'
            new_order = Order(id = new_order_id)
            db.session.add(new_order)
            new_order.initiative_id = order.initiative_id
            now = datetime.now(tz=timezone.utc)
            new_order.products = product_list
            new_order.total = sum([product['quantity']*product['price']
                               for product in new_order.products])
            new_order.income_id = order.income_id
            new_order.cashflow_id = order.cashflow_id
            new_order.site_id = order.site_id
            new_order.status = OrderStatus.new
            new_order.create_timestamp = int(now.timestamp())
            new_order.hub_id = order.hub_id
            categories = [product.get('categoryId', -1) for product in new_order.products]
            new_order.categories = Category.query.filter(Category.id.in_(categories), Category.hub_id == current_user.hub_id).all()
            new_order.parents = [order]
            message = f'заявка получена разделением из заявки {order_id}'
            event = OrderEvent(user_id=current_user.id, order_id=new_order_id, type=EventType.splitted,
                               data=message, timestamp=datetime.now(tz=timezone.utc))
            db.session.add(event)
            db.session.commit()
            SendEmailNotification('new', new_order)
        
        
        event = OrderEvent(user_id=current_user.id, order_id=order_id, type=EventType.splitted,
                           data=message_flash, timestamp=datetime.now(tz=timezone.utc))
        db.session.add(event)
        db.session.commit()

        Order.UpdateOrdersPositions(current_user.hub_id)

        OrderLimit.update_current(current_user.hub_id)

        flash(message_flash)
            
    else:
        for error in form.products.errors:
            flash(error)
    return redirect(url_for('main.ShowIndex'))

@bp.route('/orders/duplicate/<order_id>')
@login_required
@role_required([UserRoles.admin, UserRoles.initiative, UserRoles.purchaser])
@ecwid_required
def DuplicateOrder(order_id):
    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))

    order_id = GetNewOrderNumber()
    new_order = Order(id = order_id)
    db.session.add(new_order)
    
    new_order.initiative = current_user

    now = datetime.now(tz=timezone.utc)

    new_order.products = order.products
    new_order.total = order.total
    new_order.income_id = order.income_id
    new_order.cashflow_id = order.cashflow_id
    new_order.site_id = order.site_id
    new_order.status = OrderStatus.new
    new_order.create_timestamp = int(now.timestamp())

    new_order.hub_id = current_user.hub_id
    new_order.categories = order.categories

    message = 'заявка клонирована с номером {}'.format(new_order.id)
    event = OrderEvent(user_id=current_user.id, order_id=order.id,
                       type=EventType.duplicated, data=message, timestamp=datetime.now(tz=timezone.utc))
    db.session.add(event)
    message = 'заявка клонирована из заявки {}'.format(order.id)
    event = OrderEvent(user_id=current_user.id, order_id=new_order.id,
                       type=EventType.duplicated, data=message, timestamp=datetime.now(tz=timezone.utc))
    db.session.add(event)
    db.session.commit()

    Order.UpdateOrdersPositions(current_user.hub_id)

    if order.project_id is not None and order.cashflow_id is not None:
        OrderLimit.update_current(
            current_user.hub_id,
            project_id=order.project_id,
            cashflow_id=order.cashflow_id
        )
    
    flash(f'Заявка успешно клонирована. Номер новой заявки {new_order.id}. Вы перемещены в новую заявку.')

    SendEmailNotification('new', new_order)

    return redirect(url_for('main.ShowOrder', order_id=order_id))


@bp.route('/orders/quantity/<order_id>', methods=['POST'])
@login_required
@role_required([UserRoles.admin, UserRoles.initiative, UserRoles.purchaser])
@ecwid_required
def SaveQuantity(order_id):

    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))

    if order.status == OrderStatus.approved:
        flash('Нельзя модифицировать согласованную заявку.')
        return redirect(url_for('main.ShowIndex'))

    form = ChangeQuantityForm()
    if form.validate_on_submit():
        for i, product in enumerate(order.products):
            if form.product_id.data == product['id']:
                break
        else:
            flash('Указанный товар не найден в заявке.')
            return redirect(url_for('main.ShowOrder', order_id=order_id))

        if product['quantity'] != form.product_quantity.data:
            message = '{} количество было {} стало {}'.format(
                product['sku'], product['quantity'], form.product_quantity.data)
            product['quantity'] = form.product_quantity.data
            event = OrderEvent(user_id=current_user.id, order_id=order_id,
                               type=EventType.quantity, data=message, timestamp=datetime.now(tz=timezone.utc))
            db.session.add(event)

        if form.product_measurement.data != '':
            changed = False
            old_measurement = ''
            if 'selectedOptions' in product:
                if product['selectedOptions'][0]['value'] != form.product_measurement.data:
                    old_measurement = product['selectedOptions'][0]['value']
                    product['selectedOptions'][0]['value'] = form.product_measurement.data
                    changed = True
                product['selectedOptions'][0]['name'] = 'Единица измерения'
            else:
                changed = True
                product['selectedOptions'] = [
                    {'name': 'Единица измерения', 'value': form.product_measurement.data}]
            if changed is True:
                message = '{} единицы были {} стали {}'.format(
                    product['sku'], old_measurement, form.product_measurement.data)
                event = OrderEvent(user_id=current_user.id, order_id=order_id,
                                   type=EventType.measurement, data=message, timestamp=datetime.now(tz=timezone.utc))
                db.session.add(event)

        approvals = OrderApproval.query.join(User).filter(
            OrderApproval.order_id == order_id, User.hub_id == current_user.hub_id).all()
        for approval in approvals:
            db.session.delete(approval)

        order.total = sum([p['quantity']*p['price'] for p in order.products])
        order.status = OrderStatus.modified

        flag_modified(order, 'products')

        db.session.commit()

        if order.project_id is not None and order.cashflow_id is not None:
            OrderLimit.update_current(
                current_user.hub_id,
                project_id=order.project_id,
                cashflow_id=order.cashflow_id
            )

        flash('Позиция {} была изменена.'.format(product['sku']))

    else:
        for error in form.product_id.errors + form.product_quantity.errors + form.product_measurement.errors:
            flash(error)
    return redirect(url_for('main.ShowOrder', order_id=order_id))


@bp.route('/orders/excel1/<order_id>')
@login_required
@role_forbidden([UserRoles.default])
@ecwid_required
def GetExcelReport1(order_id):
    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))

    order_products = [p for p in order.products if p['quantity'] > 0]

    data_len = len(order_products)
    starting_row = 11
    wb = load_workbook(filename='template.xlsx')
    ws = wb.active
    ws['P17'] = order.initiative.name
    if data_len > 1:
        for merged_cell in ws.merged_cells.ranges:
            if merged_cell.bounds[1] >= starting_row:
                merged_cell.shift(0, data_len)
        ws.insert_rows(starting_row, data_len-1)
    for k, i in enumerate(range(starting_row, starting_row+data_len)):
        product = order_products[k]
        ws.row_dimensions[i].height = 50
        if data_len > 1:
            for j in range(1, 20):
                target_cell = ws.cell(row=i, column=j)
                source_cell = ws.cell(
                    row=starting_row + data_len - 1, column=j)
                target_cell._style = copy(source_cell._style)
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
        ws.cell(i, 1).value = k + 1
        ws.cell(i, 5).value = product['name']
        ws.cell(i, 3).value = order.site.project.name if order.site is not None else ''
        ws.cell(i, 7).value = product.get('vendor', '')
        ws.cell(i, 8).value = product['quantity']
        c1 = ws.cell(i, 8).coordinate
        ws.cell(i, 10).value = product['price']
        c2 = ws.cell(i, 10).coordinate
        ws.cell(i, 12).value = f"={c1}*{c2}"

    data = save_virtual_workbook(wb)
    return Response(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment;filename=report.xlsx'})


@bp.route('/orders/excel2/<order_id>')
@login_required
@role_forbidden([UserRoles.default])
@ecwid_required
def GetExcelReport2(order_id):
    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))

    order_products = [p for p in order.products if p['quantity'] > 0]

    data_len = len(order_products)
    starting_row = 2
    wb = load_workbook(filename='template2.xlsx')
    ws = wb.active

    ws.title = order.site.name if order.site is not None else 'Объект не указан'

    i = starting_row
    for product in order_products:
        ws.cell(i, 1).value = product['sku']
        ws.cell(i, 2).value = product['name']
        if 'selectedOptions' in product:
            ws.cell(i, 3).value = product['selectedOptions'][0]['value']
        ws.cell(i, 4).value = product['price']
        ws.cell(i, 5).value = product['quantity']
        ws.cell(i, 6).value = product['price'] * product['quantity']
        i += 1

    data = save_virtual_workbook(wb)
    return Response(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment;filename=report.xlsx'})


@bp.route('/orders/notify/<order_id>')
@login_required
@role_required([UserRoles.admin, UserRoles.initiative, UserRoles.purchaser])
@ecwid_required
def NotifyApprovers(order_id):
    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))
    SendEmailNotification('modified', order)
    flash('Уведомление успешно выслано.')
    return redirect(url_for('main.ShowOrder', order_id=order_id))


@bp.route('/orders/dealdone/<order_id>')
@login_required
@role_required([UserRoles.admin, UserRoles.purchaser])
@ecwid_required
def SetDealDone(order_id):
    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))

    if order.dealdone is True:
        flash('Заявка уже законтрактована.')
    else:
        order.dealdone = True
        event = OrderEvent(user_id=current_user.id, order_id=order_id, type=EventType.dealdone,
                           data='заявка законтрактована', timestamp=datetime.now(tz=timezone.utc))
        db.session.add(event)
        flash('Заявка законтрактована.')
        db.session.commit()
    return redirect(url_for('main.ShowOrder', order_id=order_id))


def Prepare1CReport(order, excel_date):

    order_products = [p for p in order.products if p['quantity'] > 0]

    data_len = len(order_products)
    if data_len > 0:
        categories = Category.query.filter(
            Category.hub_id == order.initiative.hub_id).all()
        starting_row = 3
        wb = load_workbook(filename='template1C.xlsx')
        ws = wb['Заявка']
        for merged_cell in ws.merged_cells.ranges:
            if merged_cell.bounds[1] >= starting_row:
                merged_cell.shift(0, data_len)

        ws.insert_rows(starting_row, data_len)
        for k, i in enumerate(range(starting_row, starting_row+data_len)):
            product = order_products[k]
            ws.row_dimensions[i].height = 50

            for j in range(1, 32):
                target_cell = ws.cell(row=i, column=j)
                source_cell = ws.cell(row=starting_row + data_len, column=j)
                target_cell._style = copy(source_cell._style)
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            # Object

            ws.cell(i, 2).value = order.site.name if order.site is not None else ''
            # Initiative

            ws.cell(i, 5).value = order.initiative.name
            ws.cell(i, 6).value = 'Для собственных нужд'

            product_cat = product.get('categoryId', 0)
            for cat in categories:
                if product_cat == cat.id:
                    ws.cell(i, 10).value = cat.functional_budget
                    ws.cell(i, 24).value = cat.responsible
                    break
            else:
                ws.cell(i, 10).value = ''
                ws.cell(i, 24).value = ''

            ws.cell(i, 11).value = order.income_statement.name if order.income_statement is not None else ''
            ws.cell(i, 12).value = order.cashflow_statement.name if order.cashflow_statement is not None else ''
            ws.cell(i, 15).value = 'Непроектные МТР и СИЗ'

            # Measurement
            if 'selectedOptions' in product:
                ws.cell(i, 19).value = product['selectedOptions'][0]['value']
                ws.cell(i, 23).value = ', '.join([p['value'] for p in product['selectedOptions'][1:]])
            # Product Name
            ws.cell(i, 20).value = product.get('name', '')
            # Quantity
            ws.cell(i, 22).value = product.get('quantity', '')
            ws.cell(i, 29).value = excel_date

            ws.cell(i, 31).value = product.get('price', '')

            ws.cell(i, 30).value = product.get('vendor', '')
        data = save_virtual_workbook(wb)
        return data
    else:
        return None


@bp.route('/orders/excel1C/<order_id>')
@login_required
@role_required([UserRoles.admin, UserRoles.validator, UserRoles.purchaser])
@ecwid_required
def GetExcelReport1C(order_id):
    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))
    try:
        excel_date = request.args.get('date', default=date.today(
        ), type=lambda x: datetime.strptime(x, "%Y-%m-%d").date())
    except:
        excel_date = date.today()
    excel_send = request.args.get('send', default=False, type=bool)
    data = Prepare1CReport(order, excel_date)
    if data is None:
        flash('Не удалось получить выгрузку.')
        return redirect(url_for('main.ShowOrder', order_id=order_id))
    if excel_send is False:
        return Response(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment;filename=pushkind_{}.xlsx'.format(order.id)})
    else:
        app_data = AppSettings.query.filter_by(
            hub_id=current_user.hub_id).first()
        if app_data is not None and app_data.email_1C is not None:
            SendEmail1C([app_data.email_1C], order, data)
            message = f'отправлена на {app_data.email_1C}'
            event = OrderEvent(user_id=current_user.id, order_id=order.id,
                               type=EventType.exported, data=message, timestamp=datetime.now(tz=timezone.utc))
            db.session.add(event)
            order.exported = True
            db.session.commit()
            flash(f'Заявка отправлена на {app_data.email_1C}')
        else:
            flash('Email для отправки в 1С не настроен администратором.')
        return redirect(url_for('main.ShowOrder', order_id=order_id))



@bp.route('/orders/approval/<order_id>', methods=['POST'])
@login_required
@role_required([UserRoles.validator])
def SaveApproval(order_id):
    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))
    form = OrderApprovalForm()
    if form.validate_on_submit():
    
        last_status = order.status
    
        position_approval = OrderPosition.query.filter_by(
            order_id=order_id, position_id=current_user.position_id).first()
        if form.comment.data != '':
            message = form.comment.data.strip()
        else:
            message = 'без комментария'
        if form.product_id.data is None:
            OrderApproval.query.filter_by(
                order_id=order_id, user_id=current_user.id).delete()

            position_disapprovals = OrderApproval.query.filter_by(order_id=order_id).join(
                User).filter(User.position_id == current_user.position_id).all()

            for disapproval in position_disapprovals:
                db.session.delete(disapproval)

            order_approval = OrderApproval(
                order_id=order_id, product_id=None, user_id=current_user.id, remark=message)
            db.session.add(order_approval)
            event = OrderEvent(user_id=current_user.id, order_id=order_id,
                               type=EventType.approved, data=message, timestamp=datetime.now(tz=timezone.utc))
            if position_approval is not None:
                position_approval.approved = True
                position_approval.user = current_user
                position_approval.timestamp = datetime.utcnow()
        else:
            OrderApproval.query.filter_by(
                order_id=order_id, user_id=current_user.id, product_id=None).delete()
            if form.product_id.data == 0:
                event = OrderEvent(user_id=current_user.id, order_id=order_id,
                                   type=EventType.disapproved, data=message, timestamp=datetime.now(tz=timezone.utc))
                product_approval = OrderApproval(
                    order_id=order_id, product_id=0, user_id=current_user.id, remark=message)
                db.session.add(product_approval)
                if position_approval is not None:
                    position_approval.approved = False
                    position_approval.user = current_user
                    position_approval.timestamp = datetime.utcnow()
            else:
                for product in order.products:
                    if form.product_id.data == product['id']:
                        break
                else:
                    flash('Указанный позиция не найдена в заявке.')
                    return redirect(url_for('main.ShowOrder', order_id=order_id))
                product_approval = OrderApproval.query.filter_by(
                    order_id=order_id, user_id=current_user.id, product_id=form.product_id.data).first()
                if product_approval is None:
                    product_approval = OrderApproval(
                        order_id=order_id, product_id=form.product_id.data, user_id=current_user.id)
                    db.session.add(product_approval)
                product_approval.remark = message
                message = 'к позиции "{}" '.format(product['name']) + message
                event = OrderEvent(user_id=current_user.id, order_id=order_id,
                                   type=EventType.disapproved, data=message, timestamp=datetime.now(tz=timezone.utc))
                if position_approval is not None:
                    position_approval.approved = False
                    position_approval.user = current_user
                    position_approval.timestamp = datetime.utcnow()
        db.session.add(event)
        order.UpdateOrderStatus()
        db.session.commit()
        flash('Согласование сохранено.')
        
        if order.status != last_status:
            if order.status == OrderStatus.approved:
                SendEmailNotification('approved', order)
                app_data = AppSettings.query.filter_by(hub_id = current_user.hub_id).first()
                if app_data is not None and app_data.email_1C is not None and app_data.notify_1C is True:
                    data = Prepare1CReport(order, date.today() + timedelta(days = 14))
                    if data is not None:
                        SendEmail1C([app_data.email_1C], order, data)

            if order.project_id is not None and order.cashflow_id is not None:
                OrderLimit.update_current(
                    current_user.hub_id,
                    project_id=order.project_id,
                    cashflow_id=order.cashflow_id
                )

            elif order.status == OrderStatus.not_approved:
                SendEmailNotification('disapproved', order)
        
    else:
        for error in form.product_id.errors + form.comment.errors:
            flash(error)
    return redirect(url_for('main.ShowOrder', order_id=order_id))


@bp.route('/orders/statements/<order_id>', methods=['POST'])
@login_required
@role_required([UserRoles.admin, UserRoles.initiative, UserRoles.validator, UserRoles.purchaser])
@ecwid_required
def SaveStatements(order_id):
    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))
    form = ApproverForm()
    
    incomes = IncomeStatement.query.filter(IncomeStatement.hub_id == current_user.hub_id).order_by(IncomeStatement.name).all()
    cashflows = CashflowStatement.query.filter(CashflowStatement.hub_id == current_user.hub_id).order_by(CashflowStatement.name).all()
    
    form.income_statement.choices = [(i.id, i.name) for i in incomes]
    form.cashflow_statement.choices = [(c.id, c.name) for c in cashflows]   
    
    if form.validate_on_submit() is True:
    
        income = IncomeStatement.query.filter_by(id=form.income_statement.data, hub_id = current_user.hub_id).first()
        cashflow = CashflowStatement.query.filter_by(id=form.cashflow_statement.data, hub_id = current_user.hub_id).first()
        
        income_name_last = order.income_statement.name if order.income_statement is not None else 'не указана'
        cashflow_name_last = order.cashflow_statement.name if order.cashflow_statement is not None else 'не указана'
        
        if income_name_last != income.name:
            message = f'статья БДР была "{income_name_last}" стала "{income.name}"'
            order.income_statement = income
            event = OrderEvent(user_id=current_user.id, order_id=order_id,
                               type=EventType.income_statement, data=message, timestamp=datetime.now(tz=timezone.utc))
            db.session.add(event)
        if cashflow_name_last != cashflow.name:
            message = f'статья БДДС была "{cashflow_name_last}" стала "{cashflow.name}"'
            order.cashflow_statement = cashflow
            event = OrderEvent(user_id=current_user.id, order_id=order_id,
                               type=EventType.cashflow_statement, data=message, timestamp=datetime.now(tz=timezone.utc))
            db.session.add(event)
        db.session.commit()

        if order.project_id is not None and order.cashflow_id is not None:
            OrderLimit.update_current(
                current_user.hub_id,
                project_id=order.project_id,
                cashflow_id=order.cashflow_id
            )

        flash('Статьи БДДР и БДДС успешно сохранены.')

    else:
        for error in form.income_statement.errors + form.cashflow_statement.errors:
            flash(error)
    return redirect(url_for('main.ShowOrder', order_id=order_id))


@bp.route('/orders/parameters/<order_id>', methods=['POST'])
@login_required
@role_required([UserRoles.admin, UserRoles.initiative, UserRoles.validator, UserRoles.purchaser])
@ecwid_required
def SaveParameters(order_id):
    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))
    form = InitiativeForm()

    projects = Project.query.filter(
        Project.hub_id == current_user.hub_id).order_by(Project.name).all()
    categories = Category.query.filter(
        Category.hub_id == current_user.hub_id).all()

    form.categories.choices = [(c.id, c.name) for c in categories]
    form.project.choices = [(p.id, p.name) for p in projects]

    project = Project.query.filter_by(
        id=form.project.data, hub_id=current_user.hub_id).first()
    if project is not None:
        form.site.choices = [(s.id, s.name) for s in project.sites]
    else:
        form.site.choices = []

    if form.validate_on_submit() is True:
        new_site = Site.query.filter_by(
            id=form.site.data, project_id=form.project.data).first()
        if new_site is not None and (order.site is None or order.site.id != new_site.id):
            message = 'объект изменён с {} на {}'.format(
                order.site.name if order.site else '', new_site.name)
            event = OrderEvent(user_id=current_user.id, order_id=order_id,
                               type=EventType.site, data=message, timestamp=datetime.now(tz=timezone.utc))
            db.session.add(event)
            order.site = Site.query.filter_by(
                id=form.site.data, project_id=form.project.data).first()
        order.categories = Category.query.filter(Category.id.in_(
            form.categories.data), Category.hub_id == current_user.hub_id).all()
        db.session.commit()
        Order.UpdateOrdersPositions(current_user.hub_id, order_id)
        if order.project_id is not None and order.cashflow_id is not None:
            OrderLimit.update_current(
                current_user.hub_id,
                project_id=order.project_id,
                cashflow_id=order.cashflow_id
            )
        flash('Параметры заявки успешно сохранены.')
    else:
        for error in form.project.errors + form.site.errors + form.categories.errors:
            flash(error)
    return redirect(url_for('main.ShowOrder', order_id=order_id))


@bp.route('/orders/comment/<order_id>', methods=['POST'])
@login_required
@role_required([UserRoles.admin, UserRoles.initiative, UserRoles.validator, UserRoles.purchaser])
def LeaveComment(order_id):
    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))
    form = LeaveCommentForm()
    if form.validate_on_submit():
        stripped = form.comment.data.strip()
        if len(stripped) > 0:
            comment = OrderEvent(
                user_id=current_user.id,
                order_id=order_id,
                type=EventType.commented,
                data=stripped,
                timestamp=datetime.now(tz=timezone.utc)
            )
            db.session.add(comment)
            flash('Комментарий успешно добавлен.')
        else:
            flash('Комментарий не может быть пустым.')
        db.session.commit()
    return redirect(url_for('main.ShowOrder', order_id=order_id))


@bp.route('/orders/process/<order_id>')
@login_required
@role_required([UserRoles.admin, UserRoles.purchaser])
@ecwid_required
def ProcessHubOrder(order_id):
    order = GetOrder(order_id)
    if order is None:
        flash('Заявка с таким номером не найдена.')
        return redirect(url_for('main.ShowIndex'))

    template = order.to_ecwid()
    template['email'] = current_user.email
    stores = Ecwid.query.filter(Ecwid.hub_id == current_user.hub_id).all()
    got_orders = {}
    for store in stores:
        products = list()
        total = 0
        for product in template['items']:
            try:
                dash = product['sku'].index('-')
            except ValueError:
                continue
            if product['sku'][:dash] == str(store.id):
                product_new = product.copy()
                product_new['sku'] = product_new['sku'][dash+1:]
                products.append(product_new)
                total += product_new['price'] * product_new['quantity']
        if len(products) == 0:
            continue
        items = template['items']
        template['items'] = products
        template['total'] = total

        try:
            result = store.SetStoreOrder(template)
            got_orders[store.name] = result['id']
        except EcwidAPIException as e:
            flash(f'Не удалось перезаказать товары у {store.name}.')
        template['items'] = items

    if len(got_orders) > 0:
        message = ', '.join(
            f'{vendor} (#{order_id})' for vendor, order_id in got_orders.items()
        )

        event = OrderEvent(
            user_id=current_user.id,
            order_id=order_id,
            type=EventType.purchased,
            data=message,
            timestamp=datetime.now(tz=timezone.utc)
        )

        order.purchased = True

        db.session.add(event)
        db.session.commit()

        flash(f'Заявка была отправлена поставщикам: {message}')
    else:
        flash('Не удалось перезаказать данные товары у зарегистрованных поставщиков.')

    return redirect(url_for('main.ShowOrder', order_id=order_id))
