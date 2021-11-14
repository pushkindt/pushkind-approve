from datetime import datetime
from app import db
from flask_login import current_user, login_required
from app.main import bp
from app.models import OrderApproval, OrderPosition, OrderStatus, User
from app.models import UserRoles, Category, Project, Position, Order
from app.models import OrderCategory, Site
from flask import render_template, redirect, url_for, flash, Response
from app.main.forms import UserRolesForm, UserSettingsForm
from sqlalchemy import or_
from app.main.utils import role_required, role_forbidden
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

'''
################################################################################
Settings page
################################################################################
'''


def RemoveExcessivePosition():
    positions = Position.query.filter(Position.users == None).all()
    for position in positions:
        position.approvals = []
        db.session.delete(position)
    db.session.commit()


@bp.route('/settings/', methods=['GET', 'POST'])
@login_required
@role_forbidden([UserRoles.default])
def ShowSettings():

    projects = Project.query
    if current_user.role != UserRoles.admin:
        projects = projects.filter_by(enabled=True)
    projects = projects.filter_by(hub_id=current_user.hub_id)
    projects = projects.order_by(Project.name).all()

    categories = Category.query.filter(
        Category.hub_id == current_user.hub_id).all()
    if current_user.role == UserRoles.admin:
        user_form = UserRolesForm()
    else:
        user_form = UserSettingsForm()

    if current_user.role in [UserRoles.admin, UserRoles.purchaser, UserRoles.validator]:
        user_form.about_user.categories.choices = [
            (c.id, c.name) for c in categories]
        user_form.about_user.projects.choices = [
            (p.id, p.name) for p in projects]

    if user_form.submit.data:
        if user_form.validate_on_submit():
            if current_user.role == UserRoles.admin:
                user = User.query.filter(
                    User.id == user_form.user_id.data).first()
                if user is None:
                    flash('Пользователь не найден.')
                    return redirect(url_for('main.ShowSettings'))
                user.hub_id = current_user.hub_id
                user.role = user_form.role.data
                user.note = user_form.note.data
            else:
                user = current_user

            if user_form.about_user.position.data is not None and user_form.about_user.position.data != '':
                position_name = user_form.about_user.position.data.strip().lower()
                position = Position.query.filter_by(
                    name=position_name, hub_id=user.hub_id).first()
                if position is None:
                    position = Position(name=position_name, hub_id=user.hub_id)
                user.position = position
            else:
                user.position = None

            if user.role in [UserRoles.purchaser, UserRoles.validator]:
                if user_form.about_user.categories.data is not None and len(user_form.about_user.categories.data) > 0:
                    user.categories = Category.query.filter(
                        Category.id.in_(user_form.about_user.categories.data)).all()
                else:
                    user.categories = []
                if user_form.about_user.projects.data is not None and len(user_form.about_user.projects.data) > 0:
                    user.projects = Project.query.filter(
                        Project.id.in_(user_form.about_user.projects.data)).all()
                else:
                    user.projects = []
            else:
                user.categories = []
                user.projects = []

            user.phone = user_form.about_user.phone.data
            user.location = user_form.about_user.location.data
            user.email_new = user_form.about_user.email_new.data
            user.email_modified = user_form.about_user.email_modified.data
            user.email_disapproved = user_form.about_user.email_disapproved.data
            user.email_approved = user_form.about_user.email_approved.data
            user.name = user_form.about_user.full_name.data.strip()
            db.session.commit()

            RemoveExcessivePosition()

            if user.role in [UserRoles.purchaser, UserRoles.validator]:
                Order.UpdateOrdersPositions(current_user.hub_id)

            flash('Данные успешно сохранены.')
        else:
            errors_list = user_form.about_user.full_name.errors + user_form.about_user.phone.errors + \
                user_form.about_user.categories.errors + user_form.about_user.projects.errors
            for error in errors_list:
                flash(error)
        return redirect(url_for('main.ShowSettings'))

    if current_user.role == UserRoles.admin:
        users = User.query.filter(or_(User.role == UserRoles.default, User.hub_id ==
                                  current_user.hub_id)).order_by(User.name, User.email).all()
        return render_template('settings.html', user_form=user_form, users=users)
    else:
        return render_template('settings.html', user_form=user_form)


@bp.route('/users/remove/<int:user_id>')
@login_required
@role_required([UserRoles.admin])
def RemoveUser(user_id):
    user = User.query.filter(User.id == user_id, or_(
        User.role == UserRoles.default, User.hub_id == current_user.hub_id)).first()
    if user is None:
        flash('Пользователь не найден.')
        return redirect(url_for('main.ShowSettings'))

    for order in user.orders:
        order.initiative = current_user

    db.session.delete(user)
    db.session.commit()

    RemoveExcessivePosition()

    if user.role in [UserRoles.purchaser, UserRoles.validator]:
        Order.UpdateOrdersPositions(current_user.hub_id)

    flash('Пользователь успешно удалён.')
    return redirect(url_for('main.ShowSettings'))


@bp.route('/users/download')
@login_required
@role_required([UserRoles.admin])
def DownloadUsers():
    users = User.query.filter(
        or_(User.role == UserRoles.default, User.hub_id == current_user.hub_id)
    ).order_by(User.name, User.email).all()
    wb = Workbook()
    ws = wb.active
    for i, header in enumerate([
        'Имя',
        'Телефон',
        'Email',
        'Роль',
        'Площадка',
        'Права',
        'Заметка',
        'Активность',
        'Регистрация',
        'Согласованных заявок пользователя',
        'Сумма согласованных заявок пользователя',
        'Согласовал заявок',
        'Должен согласовать заявок',
        'Номер для согласования'
    ], start=1):
        ws.cell(1, i).value = header

    for i, user in enumerate(users, start=2):
        ws.cell(i, 1).value = user.name
        ws.cell(i, 2).value = user.phone
        ws.cell(i, 3).value = user.email        
        ws.cell(i, 4).value = user.position.name if user.position is not None else ''
        ws.cell(i, 5).value = user.location
        ws.cell(i, 6).value = user.role
        ws.cell(i, 7).value = user.note
        ws.cell(i, 8).value = user.last_seen
        ws.cell(i, 9).value = user.registered

        # Orders which user is initiative for
        orders = Order.query.filter_by(
            initiative_id=user.id,
            status=OrderStatus.approved
        ).all()
        ws.cell(i, 10).value = len(orders)
        ws.cell(i, 11).value = sum([o.total for o in orders])

        if user.role in [UserRoles.purchaser, UserRoles.validator]:
            # Orders approved by user
            orders = Order.query.filter_by(
                hub_id=current_user.hub_id
            ).join(OrderApproval).filter_by(
                user_id=user.id,
                product_id=None
            ).all()
            ws.cell(i, 12).value = len(orders)

            # Orders to be approved
            orders = Order.query.filter(
                Order.hub_id == current_user.hub_id,
                or_(
                    Order.status == OrderStatus.new,
                    Order.status == OrderStatus.partly_approved,
                    Order.status == OrderStatus.modified
                ),
                ~Order.user_approvals.any(OrderApproval.user_id == user.id),
                ~Order.children.any()
            )
            orders = orders.join(OrderPosition)
            orders = orders.filter_by(position_id=user.position_id)
            orders = orders.join(OrderCategory)
            orders = orders.filter(OrderCategory.category_id.in_([cat.id for cat in user.categories]))
            orders = orders.join(Site)
            orders = orders.filter(Site.project_id.in_([p.id for p in user.projects]))
            orders = orders.all()
            ws.cell(i, 13).value = len(orders)
            ws.cell(i, 14).value = ', '.join([o.id for o in orders])
        else:
            ws.cell(i, 12).value = 0
            ws.cell(i, 13).value = 0
    
    data = save_virtual_workbook(wb)
    return Response(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment;filename=users.xlsx'})